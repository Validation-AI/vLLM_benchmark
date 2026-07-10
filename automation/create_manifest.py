#!/usr/bin/env python3
"""
One-time helper: generates automation_v0.xlsx with the correct schema and test rows.

Run this whenever you need to bootstrap or reset the manifest:
    python3 automation/create_manifest.py

The script is safe to re-run: it overwrites the output file but does NOT touch
any result columns (last_run_at, last_status, last_build_number) that may have
been written back by a previous run — those only exist once writeback_results.py
has been used.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path(__file__).parent / "manifests" / "serving_tuning" / "automation_v0.xlsx"
SHEET_NAME = "serving_tuning"

# ---------------------------------------------------------------------------
# Column schema
# ---------------------------------------------------------------------------
# "input"  columns are read by read_excel_manifest.py / the pipeline
# "output" columns are written back by writeback_results.py after each run
COLUMNS = [
    # --- identity ----------------------------------------------------------
    ("enabled",                  "input",  "TRUE = include in runs, FALSE = skip"),
    # --- model -------------------------------------------------------------
    ("model_id",                 "input",  "HuggingFace model path, e.g. meta-llama/Llama-2-7b-hf"),
    ("hugginface_path_185",      "input",  "Model cache path on GNR630185, e.g. /localdisk2"),
    ("hugginface_path_124",      "input",  "Model cache path on r12s04, e.g. /localdisk3"),
    # --- parallelism -------------------------------------------------------
    ("tp",                       "input",  "Tensor-parallel degree (1, 2, 4 …)"),
    ("pp",                       "input",  "Pipeline-parallel degree (usually 1)"),
    ("dp",                       "input",  "Data-parallel degree (usually 1; >1 enables router_dp)"),
    ("dp_mode",                  "input",  "none | router_dp | native_dp  (leave blank to auto-derive from dp)"),
    ("extra_args",               "input",  "Per-row server args appended to the job-level extra_args, e.g. --max-model-len 8192"),
    # --- write-back (filled by writeback_results.py) -----------------------
    ("last_run_at",              "output", "ISO date of last completed run, written by automation"),
    ("last_status",              "output", "PASS | SLA_NOT_MET | INFRA_ERROR | MODEL_ERROR"),
    ("last_build_number",        "output", "Jenkins build number of last run"),
    ("last_batch_size",          "output", "Optimal batch size from last run"),
    ("last_throughput",          "output", "Throughput (tok/s) from last run"),
    ("last_ttft_ms",             "output", "TTFT (ms) from last run"),
    ("last_tpot_ms",             "output", "TPOT (ms) from last run"),
    # --- notes -------------------------------------------------------------
    ("notes",                    "input",  "Free-form notes, not read by automation"),
]

HEADER_ROW = [col[0] for col in COLUMNS]

# ---------------------------------------------------------------------------
# Test data  (3 rows — replace with real models when ready)
# ---------------------------------------------------------------------------
TEST_ROWS = [
    {
        "enabled":                True,
        "model_id":               "meta-llama/Llama-2-7b-hf",
        "hugginface_path_185":    "/localdisk2",
        "hugginface_path_124":    "/localdisk3",
        "tp":                     1,
        "pp":                     1,
        "dp":                     1,
        "dp_mode":                "none",
        "notes":                  "TEST: baseline 7b single-socket",
    },
    {
        "enabled":                True,
        "model_id":               "meta-llama/Llama-2-7b-hf",
        "hugginface_path_185":    "/localdisk2",
        "hugginface_path_124":    "/localdisk3",
        "tp":                     2,
        "pp":                     1,
        "dp":                     1,
        "dp_mode":                "none",
        "notes":                  "TEST: 7b TP=2",
    },
    {
        "enabled":                True,
        "model_id":               "meta-llama/Meta-Llama-3-8B-Instruct",
        "hugginface_path_185":    "/localdisk2",
        "hugginface_path_124":    "/localdisk3",
        "tp":                     1,
        "pp":                     1,
        "dp":                     1,
        "dp_mode":                "none",
        "notes":                  "TEST: 8B Instruct baseline",
    },
    {
        "enabled":                False,
        "model_id":               "meta-llama/Meta-Llama-3-8B-Instruct",
        "hugginface_path_185":    "/localdisk2",
        "hugginface_path_124":    "/localdisk3",
        "tp":                     2,
        "pp":                     1,
        "dp":                     1,
        "dp_mode":                "none",
        "notes":                  "TEST: 8B TP=2 (disabled until tp=1 validated)",
    },
]

# ---------------------------------------------------------------------------
# Colours
# ---------------------------------------------------------------------------
HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
INPUT_FILL    = PatternFill("solid", fgColor="D9E1F2")   # light blue
OUTPUT_FILL   = PatternFill("solid", fgColor="E2EFDA")   # light green
DISABLED_FILL = PatternFill("solid", fgColor="F2F2F2")   # grey


def _style_header(ws, ncols):
    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _style_data_rows(ws, nrows, columns):
    for row_idx in range(2, nrows + 2):
        enabled_cell = ws.cell(row=row_idx, column=HEADER_ROW.index("enabled") + 1)
        is_disabled = str(enabled_cell.value).upper() in ("FALSE", "0", "NO", "N")
        for col_idx, (_col_name, col_kind, _) in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if is_disabled:
                cell.fill = DISABLED_FILL
            elif col_kind == "output":
                cell.fill = OUTPUT_FILL
            else:
                cell.fill = INPUT_FILL
            cell.alignment = Alignment(horizontal="left")


def _set_column_widths(ws, columns):
    for col_idx, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(col[0]) + 4, 18)


def _add_legend(wb, columns):
    ls = wb.create_sheet("legend")
    ls.append(["Column", "Kind", "Description"])
    ls["A1"].font = Font(bold=True)
    ls["B1"].font = Font(bold=True)
    ls["C1"].font = Font(bold=True)
    for col_name, kind, desc in columns:
        ls.append([col_name, kind, desc])
    ls.column_dimensions["A"].width = 28
    ls.column_dimensions["B"].width = 10
    ls.column_dimensions["C"].width = 70


def create_manifest():
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Header row
    ws.append(HEADER_ROW)
    _style_header(ws, len(COLUMNS))
    ws.row_dimensions[1].height = 36

    # Freeze header
    ws.freeze_panes = "A2"

    # Data rows — write-back columns are left blank intentionally
    output_cols = {col for col, kind, _ in COLUMNS if kind == "output"}
    for row in TEST_ROWS:
        ws.append([
            row.get(col) if col not in output_cols else None
            for col in HEADER_ROW
        ])

    _style_data_rows(ws, len(TEST_ROWS), COLUMNS)
    _set_column_widths(ws, COLUMNS)
    _add_legend(wb, COLUMNS)

    wb.save(OUTPUT_PATH)
    print(f"Created manifest: {OUTPUT_PATH}  ({len(TEST_ROWS)} rows)")
    print(f"Input columns:  {sum(1 for _, k, _ in COLUMNS if k == 'input')}")
    print(f"Output columns: {sum(1 for _, k, _ in COLUMNS if k == 'output')}")


if __name__ == "__main__":
    create_manifest()
