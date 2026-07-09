#!/usr/bin/env python3
"""
Generate a ww20-style performance comparison xlsx after each pipeline run.

Layout mirrors the "ww20 (vllm 0.20.0)" sheet in the reference spreadsheet:
  A: Model   B: Weight   C: Precision
  D-H: Old vLLM (config, concurrency, TTFT, TPOT, throughput)
  I-M: Old SGLang reference (config, concurrency, TTFT, TPOT, throughput)
  N-Q: New vLLM (concurrency, TTFT, TPOT, throughput)
  R: Throughput ratio  New vLLM / Old vLLM
  S: Throughput ratio  New vLLM / Old SGLang
  T: Jenkins artifact link

Data sources:
  - "Old vLLM": manifest xlsx last_* columns (read BEFORE writeback updates them)
  - "New vLLM": logs/row_results/*.json from the current run
  - "SGLang":   ww20 sheet of the reference spreadsheet (static copy)
  - Weight/Precision: reference spreadsheet ww20 sheet
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def config_str(tp, pp, dp):
    """Derive human-readable config string from tp/pp/dp, e.g. DP6, TP4, TP2DP2."""
    tp = int(tp) if tp else 1
    pp = int(pp) if pp else 1
    dp = int(dp) if dp else 1
    parts = []
    if tp > 1:
        parts.append(f"TP{tp}")
    if pp > 1:
        parts.append(f"PP{pp}")
    if dp > 1:
        parts.append(f"DP{dp}")
    return "".join(parts) if parts else "DP1"


def safe_num(v):
    """Return a numeric value or None."""
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# data loaders
# ---------------------------------------------------------------------------

def load_new_results(results_dir):
    """Load current run results keyed by model_id."""
    results = {}
    for path in sorted(Path(results_dir).glob("*.json")):
        with path.open("r", encoding="utf-8") as f:
            row = json.load(f)
        mid = row.get("model_id", "")
        results[mid] = row
    return results


def load_manifest_old(manifest_xlsx):
    """Load 'old' baseline from manifest xlsx last_* columns (before writeback)."""
    wb = load_workbook(manifest_xlsx, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {h: i for i, h in enumerate(headers) if h}

    rows = {}
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        mid = vals[col.get("model_id", 1)]
        if not mid:
            continue
        rows[mid] = {
            "tp": vals[col.get("tp", -1)] if "tp" in col else None,
            "pp": vals[col.get("pp", -1)] if "pp" in col else None,
            "dp": vals[col.get("dp", -1)] if "dp" in col else None,
            "concurrency": vals[col.get("last_batch_size", -1)] if "last_batch_size" in col else None,
            "throughput": vals[col.get("last_throughput", -1)] if "last_throughput" in col else None,
            "ttft_ms": vals[col.get("last_ttft_ms", -1)] if "last_ttft_ms" in col else None,
            "tpot_ms": vals[col.get("last_tpot_ms", -1)] if "last_tpot_ms" in col else None,
        }
    wb.close()
    return rows


def load_sglang_reference(ref_xlsx, sheet_name):
    """
    Load SGLang reference data + model metadata from the reference spreadsheet.
    Returns (sglang_dict, meta_dict) keyed by model_id.
    """
    wb = load_workbook(ref_xlsx, data_only=True)
    ws = wb[sheet_name]

    sglang = {}
    meta = {}
    for r in range(8, ws.max_row + 1):
        mid = ws.cell(r, 1).value
        if not mid:
            continue
        # Model metadata (cols B, C)
        meta[mid] = {
            "weight": ws.cell(r, 2).value,
            "precision": ws.cell(r, 3).value,
        }
        # SGLang reference (cols I-M = 9-13)
        sglang[mid] = {
            "config": ws.cell(r, 9).value,
            "concurrency": ws.cell(r, 10).value,
            "ttft": ws.cell(r, 11).value,
            "tpot_ms": ws.cell(r, 12).value,
            "throughput": ws.cell(r, 13).value,
        }
    wb.close()
    return sglang, meta


# ---------------------------------------------------------------------------
# xlsx writer
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=10)
INFRA_FONT = Font(color="FF0000")


def write_report(
    new_results,
    old_data,
    sglang_data,
    model_meta,
    output_path,
    docker_name,
    build_number,
    old_build_label,
    job_url,
):
    wb = Workbook()
    ws = wb.active
    ws.title = docker_name[:31]  # sheet name max 31 chars

    # ----- row 1-3: title / config info -----
    ws["A1"] = "SLA TTFT < 5s    TPOT< 100ms"
    ws["B1"] = "1k/1k"
    ws["C1"] = "New: vLLM build {}".format(build_number)
    ws["D1"] = "configs:"
    ws["E1"] = (
        "vLLM {} vs build {}: "
        "DP6 = 6 instances in parallel; TP4 = TP4/single instance"
    ).format(old_build_label, build_number)
    ws["E2"] = "SGLang: DP6 = TP1/instance 6 instances; TP3DP1 = TP3/instance 1 instance"
    ws["E3"] = "Note: SGLang TTFT in source had mixed units (some ms, some s). Values shown as-is."

    # ----- row 6: section headers (merged cells) -----
    section_headers = [
        (1, 1, "Model"),
        (2, 2, "Weight"),
        (3, 3, "Precision"),
        (4, 8, "Old vLLM ({})".format(old_build_label)),
        (9, 13, "Old SGLang (reference)"),
        (14, 17, "New vLLM (build {})".format(build_number)),
        (18, 19, "Throughput Ratios"),
    ]
    for col_start, col_end, title in section_headers:
        cell = ws.cell(row=6, column=col_start, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        if col_start != col_end:
            ws.merge_cells(
                start_row=6, start_column=col_start,
                end_row=6, end_column=col_end,
            )
            for c in range(col_start + 1, col_end + 1):
                ws.cell(row=6, column=c).fill = HEADER_FILL

    # ----- row 7: column sub-headers -----
    sub_headers = [
        "",  # A (Model)
        "",  # B (Weight)
        "",  # C (Precision)
        "TP/PP/DP",        # D
        "Concurrency",     # E
        "TTFT (s)",        # F
        "TPOT (ms)",       # G
        "Throughput",      # H
        "TP/PP/DP",        # I
        "Concurrency",     # J
        "TTFT",            # K
        "TPOT (ms)",       # L
        "Throughput",      # M
        "Concurrency",     # N
        "TTFT (s)",        # O
        "TPOT (ms)",       # P
        "Throughput",      # Q
        "New vLLM ({}) / Old vLLM ({})".format(build_number, old_build_label),  # R
        "New vLLM ({}) / Old SGLang".format(build_number),  # S
        "",  # T (link)
    ]
    for i, hdr in enumerate(sub_headers, start=1):
        cell = ws.cell(row=7, column=i, value=hdr)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # ----- data rows starting at row 8 -----
    # Build ordered model list: all models from old manifest + any new-only models
    all_models = list(old_data.keys())
    for mid in new_results:
        if mid not in all_models:
            all_models.append(mid)

    for idx, mid in enumerate(all_models):
        row = 8 + idx
        old = old_data.get(mid, {})
        new = new_results.get(mid, {})
        sg = sglang_data.get(mid, {})
        mm = model_meta.get(mid, {})

        # A: Model
        ws.cell(row=row, column=1, value=mid)
        # B: Weight
        ws.cell(row=row, column=2, value=mm.get("weight", ""))
        # C: Precision
        ws.cell(row=row, column=3, value=mm.get("precision", ""))

        # D-H: Old vLLM
        old_tp = old.get("tp")
        old_pp = old.get("pp")
        old_dp = old.get("dp")
        if old_tp is not None:
            ws.cell(row=row, column=4, value=config_str(old_tp, old_pp, old_dp))
        ws.cell(row=row, column=5, value=old.get("concurrency"))
        old_ttft = safe_num(old.get("ttft_ms"))
        if old_ttft is not None:
            ws.cell(row=row, column=6, value=round(old_ttft / 1000, 5))  # ms -> s
        ws.cell(row=row, column=7, value=old.get("tpot_ms"))
        ws.cell(row=row, column=8, value=old.get("throughput"))

        # I-M: SGLang reference
        ws.cell(row=row, column=9, value=sg.get("config"))
        ws.cell(row=row, column=10, value=sg.get("concurrency"))
        ws.cell(row=row, column=11, value=sg.get("ttft"))
        ws.cell(row=row, column=12, value=sg.get("tpot_ms"))
        ws.cell(row=row, column=13, value=sg.get("throughput"))

        # N-Q: New vLLM
        new_status = new.get("status", "")
        if new_status in ("INFRA_ERROR", "MODEL_ERROR"):
            cell_n = ws.cell(row=row, column=14, value=new_status)
            cell_n.font = INFRA_FONT
            ws.cell(row=row, column=15, value=None)
            ws.cell(row=row, column=16, value=None)
            ws.cell(row=row, column=17, value=None)
        elif new:
            ws.cell(row=row, column=14, value=new.get("best_batch_size"))
            new_ttft = safe_num(new.get("ttft_ms"))
            if new_ttft is not None:
                ws.cell(row=row, column=15, value=round(new_ttft / 1000, 5))  # ms -> s
            ws.cell(row=row, column=16, value=new.get("tpot_ms"))
            ws.cell(row=row, column=17, value=new.get("throughput"))

        # R: New vLLM / Old vLLM throughput ratio (IFERROR formula)
        q = "Q{}".format(row)
        h = "H{}".format(row)
        ws.cell(row=row, column=18).value = \
            '=IFERROR(IF(AND(ISNUMBER({q}),ISNUMBER({h}),{h}<>0),{q}/{h},""),"")'.format(q=q, h=h)

        # S: New vLLM / Old SGLang throughput ratio (IFERROR formula)
        m = "M{}".format(row)
        ws.cell(row=row, column=19).value = \
            '=IFERROR(IF(AND(ISNUMBER({q}),ISNUMBER({m}),{m}<>0),{q}/{m},""),"")'.format(q=q, m=m)

        # T: Jenkins artifact link
        build_url = new.get("build_url", "")
        if not build_url and job_url and build_number:
            build_url = "{}/{}/".format(job_url.rstrip("/"), build_number)
        if build_url:
            ws.cell(row=row, column=20, value=build_url)

    # ----- column widths -----
    col_widths = {
        1: 45, 2: 8, 3: 10, 4: 12, 5: 12, 6: 12, 7: 12, 8: 14,
        9: 12, 10: 12, 11: 12, 12: 12, 13: 14,
        14: 12, 15: 12, 16: 12, 17: 14, 18: 22, 19: 22, 20: 50,
    }
    for c, w in col_widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # ----- number format for ratio columns -----
    for r in range(8, 8 + len(all_models)):
        for c in (18, 19):
            ws.cell(row=r, column=c).number_format = '0.00%'

    wb.save(output_path)
    print("Performance report saved: {}  ({} models)".format(output_path, len(all_models)))


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    p = argparse.ArgumentParser(
        description="Generate ww20-style performance comparison xlsx."
    )
    p.add_argument("--results-dir", required=True,
                   help="Directory containing per-row result JSON files (logs/row_results)")
    p.add_argument("--manifest-xlsx", required=True,
                   help="Manifest xlsx with old baseline in last_* columns (read BEFORE writeback)")
    p.add_argument("--reference-xlsx", required=True,
                   help="Reference spreadsheet with SGLang data and model metadata")
    p.add_argument("--reference-sheet", default="ww20 (vllm 0.20.0)",
                   help="Sheet name in reference xlsx to read SGLang/meta from")
    p.add_argument("--output-xlsx", required=True,
                   help="Output xlsx path")
    p.add_argument("--docker-name", default="vllm-cpu",
                   help="Docker image name used as sheet title (e.g. vllm-cpu-0.21.0)")
    p.add_argument("--build-number", default="",
                   help="Current Jenkins build number")
    p.add_argument("--old-build-label", default="prev",
                   help="Label for the old/baseline build (e.g. 'build 28')")
    p.add_argument("--job-url", default="",
                   help="Jenkins JOB_URL for constructing artifact links")
    args = p.parse_args()

    new_results = load_new_results(args.results_dir)
    old_data = load_manifest_old(args.manifest_xlsx)
    sglang_data, model_meta = load_sglang_reference(
        args.reference_xlsx, args.reference_sheet
    )

    write_report(
        new_results=new_results,
        old_data=old_data,
        sglang_data=sglang_data,
        model_meta=model_meta,
        output_path=args.output_xlsx,
        docker_name=args.docker_name,
        build_number=args.build_number,
        old_build_label=args.old_build_label,
        job_url=args.job_url,
    )


if __name__ == "__main__":
    main()
