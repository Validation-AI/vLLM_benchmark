#!/usr/bin/env python3
import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


def normalize_bool(value):
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y"}


def normalize_int(value, default):
    if value is None or value == "":
        return default
    return int(value)


def normalize_str(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_parallel_spec(spec):
    tokens = dict(re.findall(r"(TP|PP|DP)(\d+)", str(spec).upper()))
    return {
        "tp": int(tokens.get("TP", 1)),
        "pp": int(tokens.get("PP", 1)),
        "dp": int(tokens.get("DP", 1)),
    }


def derive_row_id(model_id, tp, pp, dp):
    """Compute a stable unique key from model identity + parallelism config."""
    return f"{model_id}__tp{tp}_pp{pp}_dp{dp}"


def row_to_dict(headers, values):
    return {headers[idx]: values[idx] for idx in range(len(headers))}


HISTORY_FIELDS = (
    "last_run_at",
    "last_status",
    "last_build_number",
    "last_batch_size",
    "last_throughput",
    "last_ttft_ms",
    "last_tpot_ms",
)


def resolve_row(raw):
    model_id = normalize_str(raw.get("model_id"))

    if not model_id:
        raise ValueError("missing model_id")

    tp = raw.get("tp")
    pp = raw.get("pp")
    dp = raw.get("dp")

    if tp in (None, "") or pp in (None, "") or dp in (None, ""):
        parsed = parse_parallel_spec(raw.get("parallel_spec") or "")
        tp = normalize_int(tp, parsed["tp"])
        pp = normalize_int(pp, parsed["pp"])
        dp = normalize_int(dp, parsed["dp"])
    else:
        tp = normalize_int(tp, 1)
        pp = normalize_int(pp, 1)
        dp = normalize_int(dp, 1)

    dp_mode = normalize_str(raw.get("dp_mode"))
    if not dp_mode:
        dp_mode = "router_dp" if dp > 1 else "none"

    hugginface_path_185 = normalize_str(raw.get("hugginface_path_185"))
    hugginface_path_124 = normalize_str(raw.get("hugginface_path_124"))
    hugginface_path = normalize_str(raw.get("hugginface_path"))
    if not hugginface_path and not hugginface_path_185 and not hugginface_path_124:
        raise ValueError(f"{model_id}: no model path found (set hugginface_path_185 and/or hugginface_path_124)")

    row_id = derive_row_id(model_id, tp, pp, dp)

    return {
        "row_id": row_id,
        "enabled": normalize_bool(raw.get("enabled", True)),
        "model_id": model_id,
        "hugginface_path": hugginface_path,
        "hugginface_path_185": hugginface_path_185,
        "hugginface_path_124": hugginface_path_124,
        "tp": tp,
        "pp": pp,
        "dp": dp,
        "dp_mode": dp_mode,
        "extra_args": normalize_str(raw.get("extra_args")),
        "last_run_at": normalize_str(raw.get("last_run_at")),
        "last_status": normalize_str(raw.get("last_status")),
        "last_build_number": normalize_str(raw.get("last_build_number")),
        "last_batch_size": normalize_str(raw.get("last_batch_size")),
        "last_throughput": normalize_str(raw.get("last_throughput")),
        "last_ttft_ms": normalize_str(raw.get("last_ttft_ms")),
        "last_tpot_ms": normalize_str(raw.get("last_tpot_ms")),
        "notes": raw.get("notes"),
        "source_sheet": raw.get("source_sheet"),
        "source_row": raw.get("source_row"),
    }


def load_state_history(state_file):
    if not state_file:
        return {}
    path = Path(state_file)
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    return data if isinstance(data, dict) else {}


def merge_row_history(row, state_history):
    merged = dict(row)
    source = "workbook" if any(row.get(field) not in (None, "") for field in HISTORY_FIELDS) else "none"
    state_entry = state_history.get(row["row_id"], {})
    if isinstance(state_entry, dict):
        for field in HISTORY_FIELDS:
            value = normalize_str(state_entry.get(field))
            if value:
                merged[field] = value
                source = "state"
    merged["history_source"] = source
    return merged


def parse_args():
    parser = argparse.ArgumentParser(description="Read serving_tuning workbook rows into JSON.")
    parser.add_argument("--xlsx", required=True, help="Path to workbook")
    parser.add_argument("--sheet", default="serving_tuning", help="Worksheet name")
    parser.add_argument("--output", required=True, help="Output JSON path")
    parser.add_argument("--row-filter", default="", help="Comma-separated model_id list")
    parser.add_argument("--include-disabled", action="store_true", help="Include disabled rows")
    parser.add_argument(
        "--state-file",
        default="",
        help="Optional state JSON whose last_* fields override workbook history.",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    xlsx_path = Path(args.xlsx)
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[args.sheet]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    state_history = load_state_history(args.state_file.strip())
    allowed_models = {item.strip() for item in args.row_filter.split(",") if item.strip()}
    rows = []

    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and value != "" for value in values):
            continue
        raw = row_to_dict(headers, values)
        row = resolve_row(raw)
        row = merge_row_history(row, state_history)
        if allowed_models and row["model_id"] not in allowed_models:
            continue
        if not args.include_disabled and not row["enabled"]:
            continue
        rows.append(row)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(rows, f, indent=2, ensure_ascii=True)

    print(f"Wrote {len(rows)} rows to {output_path}")


if __name__ == "__main__":
    main()
