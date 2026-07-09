#!/usr/bin/env python3
"""
writeback_results.py — after a Jenkins run completes, read the compiled results
and write them back into the manifest xlsx and into the state file used by
trigger_jenkins.py for staleness checks.

Usage
-----
# After a run, point at the results CSV and the manifest:
    python3 automation/writeback_results.py \
        --results-csv  logs/serving_tuning_results.csv \
        --xlsx         automation/manifests/serving_tuning/automation_v0.xlsx \
        --build-number 156 \
        --build-url    http://jenkins.example.com/job/JOB/156/

# The script is idempotent: re-running with the same build number is safe.
# It picks the best (PASS-preferred) result per row_id across all length_configs.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from serving_tuning_layout import (
    best_runs_dir,
    latest_state_path,
    legacy_runs_dir,
    legacy_state_path,
    workspace_root,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

STATUS_RANK = {"PASS": 0, "SLA_NOT_MET": 1, "MODEL_ERROR": 2, "INFRA_ERROR": 3}


def best_result_per_row(csv_path: Path) -> dict[str, dict]:
    """
    Read compiled results CSV and return one representative result per row_id.
    Priority: PASS > SLA_NOT_MET > MODEL_ERROR > INFRA_ERROR.
    Within the same status, pick the result with the highest throughput.
    """
    rows: dict[str, list[dict]] = {}
    with csv_path.open(newline="", encoding="utf-8") as f:
        for rec in csv.DictReader(f):
            rid = rec.get("row_id", "").strip()
            if rid:
                rows.setdefault(rid, []).append(rec)

    best: dict[str, dict] = {}
    for rid, candidates in rows.items():
        def sort_key(r):
            rank = STATUS_RANK.get(r.get("status", ""), 99)
            try:
                thp = float(r.get("throughput") or 0)
            except ValueError:
                thp = 0.0
            return (rank, -thp)
        best[rid] = sorted(candidates, key=sort_key)[0]
    return best


# ---------------------------------------------------------------------------
# Manifest xlsx write-back
# ---------------------------------------------------------------------------

WRITEBACK_COLUMNS = {
    "last_run_at":     None,   # filled with today
    "last_status":     "status",
    "last_build_number": None,  # filled with --build-number arg
    "last_batch_size": "best_batch_size",
    "last_throughput": "throughput",
    "last_ttft_ms":    "ttft_ms",
    "last_tpot_ms":    "tpot_ms",
}


def _derive_row_id(model_id: str, tp, pp, dp) -> str:
    """Must match derive_row_id() in read_excel_manifest.py."""
    return f"{model_id}__tp{tp}_pp{pp}_dp{dp}"


def writeback_xlsx(xlsx_path: Path, results: dict[str, dict], build_number: str, today_str: str):
    wb = load_workbook(xlsx_path)
    if "serving_tuning" not in wb.sheetnames:
        print(f"WARNING: sheet 'serving_tuning' not found in {xlsx_path}, skipping xlsx write-back")
        return

    ws = wb["serving_tuning"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_index = {h: i + 1 for i, h in enumerate(headers) if h}

    updated = 0
    for row in ws.iter_rows(min_row=2):
        # Derive the key from model_id + tp + pp + dp columns
        def _cell(col_name):
            idx = col_index.get(col_name)
            return str(row[idx - 1].value or "").strip() if idx else ""

        model_id = _cell("model_id")
        if not model_id:
            continue
        row_id = _derive_row_id(model_id, _cell("tp"), _cell("pp"), _cell("dp"))
        if row_id not in results:
            continue

        result = results[row_id]
        first_cell = row[0]

        def _set(col_name: str, value):
            if col_name in col_index:
                ws.cell(row=first_cell.row, column=col_index[col_name]).value = value

        _set("last_run_at",       today_str)
        _set("last_status",       result.get("status", ""))
        _set("last_build_number", build_number)
        _set("last_batch_size",   result.get("best_batch_size", ""))
        _set("last_throughput",   result.get("throughput", ""))
        _set("last_ttft_ms",      result.get("ttft_ms", ""))
        _set("last_tpot_ms",      result.get("tpot_ms", ""))
        updated += 1

    wb.save(xlsx_path)
    print(f"  Wrote back {updated} rows into {xlsx_path}")


# ---------------------------------------------------------------------------
# State file (used by trigger_jenkins.py for staleness checks)
# ---------------------------------------------------------------------------

def update_state_file(state_path: Path, results: dict[str, dict], build_number: str, build_url: str, today_str: str):
    state: dict = {}
    if state_path.exists():
        with state_path.open() as f:
            state = json.load(f)

    for row_id, result in results.items():
        state[row_id] = {
            "row_id": row_id,
            "last_run_at":      today_str,
            "last_status":      result.get("status", ""),
            "last_build_number": build_number,
            "last_build_url":   build_url,
            "last_batch_size":  result.get("best_batch_size", ""),
            "last_throughput":  result.get("throughput", ""),
            "last_ttft_ms":     result.get("ttft_ms", ""),
            "last_tpot_ms":     result.get("tpot_ms", ""),
        }

    state_path.parent.mkdir(parents=True, exist_ok=True)
    with state_path.open("w") as f:
        json.dump(state, f, indent=2)
    print(f"  State file updated: {state_path}  ({len(results)} rows)")


# ---------------------------------------------------------------------------
# Runs snapshot
# ---------------------------------------------------------------------------

def save_runs_snapshot(runs_dir: Path, results: dict[str, dict], build_number: str, build_url: str, today_str: str):
    runs_dir.mkdir(parents=True, exist_ok=True)
    snapshot = {
        "build_number": build_number,
        "build_url":    build_url,
        "run_date":     today_str,
        "results":      results,
    }
    path = runs_dir / f"{today_str}_build{build_number}.json"
    with path.open("w") as f:
        json.dump(snapshot, f, indent=2)
    print(f"  Runs snapshot saved: {path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def parse_args():
    p = argparse.ArgumentParser(description="Write Jenkins run results back into the manifest.")
    p.add_argument("--results-csv",   required=True,  help="Path to logs/serving_tuning_results.csv")
    p.add_argument("--xlsx",          required=True,  help="Path to automation_v0.xlsx")
    p.add_argument("--build-number",  required=True,  help="Jenkins build number")
    p.add_argument("--build-url",     default="",     help="Jenkins build URL")
    p.add_argument("--state-file",    default="",     help="Optional compatibility mirror for the legacy state file path")
    p.add_argument("--runs-dir",      default="",     help="Optional compatibility mirror for the legacy runs/ snapshot directory")
    return p.parse_args()


def main():
    args = parse_args()
    csv_path  = Path(args.results_csv)
    xlsx_path = Path(args.xlsx)
    today_str = date.today().isoformat()
    build_number = str(args.build_number)
    build_url    = args.build_url

    if not csv_path.exists():
        sys.exit(f"ERROR: results CSV not found: {csv_path}")
    if not xlsx_path.exists():
        sys.exit(f"ERROR: manifest not found: {xlsx_path}")

    workspace = workspace_root()
    canonical_state_path = latest_state_path(workspace)
    compatibility_state_path = Path(args.state_file) if args.state_file else legacy_state_path(workspace)
    canonical_runs_dir = best_runs_dir(workspace)
    compatibility_runs_dir = Path(args.runs_dir) if args.runs_dir else legacy_runs_dir(workspace)

    print(f"Reading results from: {csv_path}")
    results = best_result_per_row(csv_path)
    print(f"  {len(results)} unique row_ids in results")

    if not results:
        print("No results found — nothing to write back.")
        return

    for row_id, r in results.items():
        print(f"  {row_id}: status={r.get('status')}  batch={r.get('best_batch_size')}  "
              f"thp={r.get('throughput')}  ttft={r.get('ttft_ms')}  tpot={r.get('tpot_ms')}")

    print(f"\nUpdating manifest xlsx: {xlsx_path}")
    writeback_xlsx(xlsx_path, results, build_number, today_str)

    state_targets = []
    for candidate in (canonical_state_path, compatibility_state_path):
        if candidate not in state_targets:
            state_targets.append(candidate)
    for state_target in state_targets:
        print(f"\nUpdating state file: {state_target}")
        update_state_file(state_target, results, build_number, build_url, today_str)

    runs_targets = []
    for candidate in (canonical_runs_dir, compatibility_runs_dir):
        if candidate not in runs_targets:
            runs_targets.append(candidate)
    for runs_target in runs_targets:
        print(f"\nSaving runs snapshot: {runs_target}")
        save_runs_snapshot(runs_target, results, build_number, build_url, today_str)

    print("\nWrite-back complete.")


if __name__ == "__main__":
    main()
