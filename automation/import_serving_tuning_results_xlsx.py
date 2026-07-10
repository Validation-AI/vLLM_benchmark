#!/usr/bin/env python3
"""
Import a compiled serving_tuning results workbook into the persistent history
layout and legacy compatibility JSON files.
"""

from __future__ import annotations

import argparse
import csv
import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from serving_tuning_layout import (
    best_runs_dir,
    full_runs_dir,
    latest_state_path,
    legacy_runs_dir,
    legacy_state_path,
    report_csv_path,
    report_xlsx_path,
    workspace_root,
)


STATUS_RANK = {"PASS": 0, "SLA_NOT_MET": 1, "MODEL_ERROR": 2, "INFRA_ERROR": 3}


def parse_args():
    parser = argparse.ArgumentParser(
        description="Import serving_tuning_results.xlsx into history/state JSON files."
    )
    parser.add_argument("--results-xlsx", required=True, help="Path to the compiled results workbook")
    parser.add_argument("--sheet", default="serving_tuning_results", help="Worksheet to import")
    parser.add_argument("--run-date", default="", help="Override run date in YYYY-MM-DD")
    parser.add_argument("--build-number", default="", help="Override build number")
    parser.add_argument("--build-url", default="", help="Override build URL")
    parser.add_argument("--state-file", default="", help="Optional compatibility mirror for the legacy state path")
    parser.add_argument("--runs-dir", default="", help="Optional compatibility mirror for the legacy runs path")
    return parser.parse_args()


def load_rows(xlsx_path: Path, sheet_name: str) -> tuple[list[str], list[dict]]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    values = list(ws.iter_rows(values_only=True))
    if not values:
        return [], []
    headers = [str(cell).strip() if cell is not None else "" for cell in values[0]]
    rows = []
    for raw in values[1:]:
        rec = {headers[i]: raw[i] for i in range(len(headers))}
        row_id = str(rec.get("row_id") or "").strip()
        if not row_id:
            continue
        rows.append(rec)
    return headers, rows


def resolve_run_date(xlsx_path: Path, requested: str) -> str:
    if requested:
        return requested
    wb = load_workbook(xlsx_path, read_only=True)
    modified = wb.properties.modified
    if modified:
        return modified.date().isoformat()
    return date.fromtimestamp(xlsx_path.stat().st_mtime).isoformat()


def resolve_unique_value(rows: list[dict], key: str, default: str = "") -> str:
    values = {
        str(row.get(key)).strip()
        for row in rows
        if row.get(key) not in (None, "") and str(row.get(key)).strip()
    }
    if len(values) == 1:
        return next(iter(values))
    return default


def best_results_per_row(rows: list[dict]) -> dict[str, dict]:
    grouped: dict[str, list[dict]] = {}
    for row in rows:
        row_id = str(row.get("row_id") or "").strip()
        if row_id:
            grouped.setdefault(row_id, []).append(row)

    best: dict[str, dict] = {}
    for row_id, candidates in grouped.items():
        def sort_key(rec: dict):
            rank = STATUS_RANK.get(str(rec.get("status") or "").strip(), 99)
            try:
                thp = float(rec.get("throughput") or 0)
            except (TypeError, ValueError):
                thp = 0.0
            return (rank, -thp)

        best[row_id] = dict(sorted(candidates, key=sort_key)[0])
    return best


def normalize_state_entry(row_id: str, result: dict, build_number: str, build_url: str, run_date: str) -> dict:
    return {
        "row_id": row_id,
        "last_run_at": run_date,
        "last_status": str(result.get("status") or ""),
        "last_build_number": build_number,
        "last_build_url": build_url,
        "last_batch_size": str(result.get("best_batch_size") or ""),
        "last_throughput": str(result.get("throughput") or ""),
        "last_ttft_ms": str(result.get("ttft_ms") or ""),
        "last_tpot_ms": str(result.get("tpot_ms") or ""),
    }


def update_state_file(state_path: Path, best_results: dict[str, dict], build_number: str, build_url: str, run_date: str):
    state = {}
    if state_path.exists():
        with state_path.open(encoding="utf-8") as handle:
            loaded = json.load(handle)
        if isinstance(loaded, dict):
            state = loaded

    for row_id, result in best_results.items():
        state[row_id] = normalize_state_entry(row_id, result, build_number, build_url, run_date)

    state_path.parent.mkdir(parents=True, exist_ok=True)
    with state_path.open("w", encoding="utf-8") as handle:
        json.dump(state, handle, indent=2)
    print(f"State file updated: {state_path} ({len(best_results)} rows)")


def save_snapshot(path: Path, payload: dict):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2)
    print(f"Snapshot saved: {path}")


def write_report_csv(headers: list[str], rows: list[dict], csv_path: Path):
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    print(f"Report saved: {csv_path}")


def copy_report_xlsx(src: Path, dest: Path):
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(src.read_bytes())
    print(f"Report saved: {dest}")


def main():
    args = parse_args()
    workspace = workspace_root()
    xlsx_path = Path(args.results_xlsx)
    if not xlsx_path.is_absolute():
        xlsx_path = workspace / xlsx_path
    if not xlsx_path.exists():
        raise SystemExit(f"ERROR: results workbook not found: {xlsx_path}")

    headers, rows = load_rows(xlsx_path, args.sheet)
    if not rows:
        raise SystemExit(f"ERROR: no result rows found in {xlsx_path}")

    run_date = resolve_run_date(xlsx_path, args.run_date)
    build_number = args.build_number or resolve_unique_value(rows, "build_number", "unknown")
    build_url = args.build_url or resolve_unique_value(rows, "build_url", "")
    best_results = best_results_per_row(rows)

    report_xlsx = report_xlsx_path(workspace, run_date, build_number)
    report_csv = report_csv_path(workspace, run_date, build_number)
    copy_report_xlsx(xlsx_path, report_xlsx)
    write_report_csv(headers, rows, report_csv)

    full_payload = {
        "build_number": build_number,
        "build_url": build_url,
        "run_date": run_date,
        "compiled_at": datetime.now().isoformat(),
        "results": rows,
    }
    best_payload = {
        "build_number": build_number,
        "build_url": build_url,
        "run_date": run_date,
        "results": best_results,
    }
    filename = f"{run_date}_build{build_number}.json"
    save_snapshot(full_runs_dir(workspace) / filename, full_payload)
    save_snapshot(best_runs_dir(workspace) / filename, best_payload)

    compatibility_runs_dir = Path(args.runs_dir) if args.runs_dir else legacy_runs_dir(workspace)
    if not compatibility_runs_dir.is_absolute():
        compatibility_runs_dir = workspace / compatibility_runs_dir
    save_snapshot(compatibility_runs_dir / filename, best_payload)

    canonical_state = latest_state_path(workspace)
    compatibility_state = Path(args.state_file) if args.state_file else legacy_state_path(workspace)
    if not compatibility_state.is_absolute():
        compatibility_state = workspace / compatibility_state

    state_targets = []
    for candidate in (canonical_state, compatibility_state):
        if candidate not in state_targets:
            state_targets.append(candidate)
    for state_target in state_targets:
        update_state_file(state_target, best_results, build_number, build_url, run_date)

    print("Import complete.")


if __name__ == "__main__":
    main()
